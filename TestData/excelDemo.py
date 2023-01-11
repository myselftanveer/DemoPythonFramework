import os

import openpyxl

book = openpyxl.load_workbook("D:\\pythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=1)
# print(cell.value)

# sheet.cell(row=3, column=3).value = "XYZ"
# print(sheet.cell(row=3, column=3).value)
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet["A2"].value)

for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "Invalid":
        for j in range(2, sheet.max_column + 1):  # to get columns:
            # print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
print('Absolute directoryname: ',os.path.dirname(os.path.abspath(__file__)))
